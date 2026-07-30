"""
검사 커맨드 카탈로그 관리.
- 필수(essential): 기본 활성화(enabled=True), 어떤 장비 점검이든 공통으로 쓰는 것
- 선택(optional): 기본 비활성화(enabled=False), 상황에 따라 켜서 쓰는 것
- custom: 사용자가 직접 추가한 것

로컬 파일(commands_catalog.yaml)에 저장/불러오기 — 토글·추가·삭제한 상태가 다음 실행 때도 유지됨.
"""
import os
import yaml
from core.atomic_io import dump_yaml_atomic

DEFAULT_CATALOG_PATH = "config/commands_catalog.yaml"

CATEGORY_LABELS = {"essential": "필수", "optional": "선택사항", "custom": "커스텀"}
EXCEL_HEADERS = ["순서", "카테고리", "사용여부", "커맨드", "설명"]

import pandas as pd
from core.paths import AppPaths

EXCEL_DEFAULT_PATH = str(AppPaths.app_root() / "data" / "eve" / "test" / "commands" / "commands_catalog.xlsx")

def _make_default_catalog():
    catalog = []
    if not os.path.exists(EXCEL_DEFAULT_PATH):
        print(f"경고: 기본 커맨드 엑셀 파일이 없습니다. ({EXCEL_DEFAULT_PATH}) 빈 카탈로그로 초기화합니다.")
        return catalog
        
    try:
        df = pd.read_excel(EXCEL_DEFAULT_PATH)
        for idx, row in df.iterrows():
            command_text = str(row.get("커맨드", "")).strip()
            if not command_text or command_text == "nan":
                continue
                
            enabled = bool(row.get("사용여부", True))
            raw_category = str(row.get("카테고리", "")).strip()
            desc = str(row.get("설명", "")).strip()
            if desc == "nan":
                desc = ""
                
            category = "essential" if raw_category == "필수" else ("optional" if raw_category == "선택사항" else "custom")
            
            catalog.append({
                "id": f"{category}_{idx}",
                "category": category,
                "command": command_text,
                "description": desc,
                "enabled": enabled
            })
    except Exception as e:
        print(f"엑셀 파일 읽기 실패: {e}")
        
    return catalog


def load_catalog(path=DEFAULT_CATALOG_PATH):
    if not os.path.exists(path):
        catalog = _make_default_catalog()
        save_catalog(catalog, path)
        return catalog
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f) or []


def save_catalog(catalog, path=DEFAULT_CATALOG_PATH):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    dump_yaml_atomic(catalog, path)


def toggle_command(catalog, command_id, enabled):
    for item in catalog:
        if item["id"] == command_id:
            item["enabled"] = enabled
            return True
    return False


def add_command(catalog, command_text, description=""):
    existing_ids = [item["id"] for item in catalog if item["category"] == "custom"]
    next_num = len(existing_ids)
    new_id = f"custom_{next_num}"
    while new_id in existing_ids:
        next_num += 1
        new_id = f"custom_{next_num}"
    catalog.append({
        "id": new_id, "category": "custom",
        "command": command_text, "description": description or "(설명 없음)",
        "enabled": True,
    })
    return new_id


def move_items(catalog, item_ids, target_index=None):
    """item_ids(여러 개 가능, 상대 순서 유지)를 카탈로그 전체(단일 목록) 기준
    target_index 위치로 이동. target_index=None이면 맨 끝. catalog는 in-place로 수정됨.
    카탈로그 파일을 매번 새로 읽어 처리하므로(_load_catalog) 클라이언트가 들고 있던
    stale한 DOM 상태와 무관하게 항상 디스크의 최신 목록을 기준으로 동작한다."""
    id_set = set(item_ids)
    moving = [item for item in catalog if item["id"] in id_set]
    if not moving:
        return
    remaining = [item for item in catalog if item["id"] not in id_set]
    insert_pos = len(remaining) if target_index is None else max(0, min(target_index, len(remaining)))
    remaining[insert_pos:insert_pos] = moving
    catalog[:] = remaining


def set_category(catalog, item_id, category):
    """카테고리(필수/선택/커스텀) 라벨만 바꾼다 — 통합 목록에서의 위치는 그대로 유지."""
    for item in catalog:
        if item["id"] == item_id:
            item["category"] = category
            return True
    return False


def remove_command(catalog, command_id):
    idx = next((i for i, item in enumerate(catalog) if item["id"] == command_id), None)
    if idx is None:
        return False
    del catalog[idx]
    return True


def get_enabled_commands(catalog):
    """하위호환용 — literal command 필드 그대로 반환(벤더 무관, 기존 호출부 안 건드림)."""
    return [item["command"] for item in catalog if item["enabled"]]


def export_to_excel(catalog, path):
    """카탈로그를 엑셀 한 장으로 내보냄. 열: 순서, 카테고리, 사용여부, 커맨드, 설명."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl 후 다시 시도하세요.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "commands_catalog"
    ws.append(EXCEL_HEADERS)
    for i, item in enumerate(catalog, start=1):
        ws.append([
            i,
            CATEGORY_LABELS.get(item.get("category"), item.get("category")),
            "TRUE" if item.get("enabled") else "FALSE",
            item.get("command", ""),
            item.get("description", ""),
        ])
    for col, width in zip("ABCDE", (6, 12, 10, 32, 30)):
        ws.column_dimensions[col].width = width
    wb.save(path)


def import_from_excel(path):
    """엑셀(export_to_excel 형식)을 읽어 새 catalog 리스트로 반환 — 기존 카탈로그를 통째로 대체할 때 씀."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl 후 다시 시도하세요.")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    label_to_category = {v: k for k, v in CATEGORY_LABELS.items()}
    truthy = {"true", "1", "y", "yes", "o", "사용", "예", "checked"}

    def parse_enabled(v):
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in truthy

    parsed = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[3]:
            continue
        order, category_label, enabled_raw, command, description = (list(row) + [None] * 5)[:5]
        category = label_to_category.get(str(category_label).strip(), "custom") if category_label else "custom"
        parsed.append({
            "order": order if isinstance(order, (int, float)) else len(parsed) + 1,
            "category": category,
            "enabled": parse_enabled(enabled_raw),
            "command": str(command).strip(),
            "description": str(description).strip() if description else "",
        })
    parsed.sort(key=lambda x: x["order"])

    counters = {"essential": 0, "optional": 0, "custom": 0}
    catalog = []
    for item in parsed:
        cat = item["category"]
        idx = counters[cat]
        counters[cat] += 1
        catalog.append({
            "id": f"{cat}_{idx}",
            "category": cat,
            "command": item["command"],
            "description": item["description"] or "(설명 없음)",
            "enabled": item["enabled"],
        })
    return catalog


if __name__ == "__main__":
    catalog = load_catalog()
    print(f"카탈로그 항목 {len(catalog)}개 (필수 {sum(1 for c in catalog if c['category']=='essential')}, "
          f"선택 {sum(1 for c in catalog if c['category']=='optional')}, "
          f"커스텀 {sum(1 for c in catalog if c['category']=='custom')})")
    print("\n활성화된 커맨드:")
    for cmd in get_enabled_commands(catalog):
        print(" -", cmd)
