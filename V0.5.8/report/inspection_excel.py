"""
범용 정기점검 보고서 엑셀 렌더러 (openpyxl).

첨부된 3개 레거시 보고서(대한전선 Arista / LGES 오창 / NC US1 특별점검)의 공통 구조를
"고객사·벤더에 종속되지 않는" 하나의 워크북 레이아웃으로 통합한 것이다:

    표지        고객사 + 프로파일(회차) 이름으로 제목을 조립 + 담당자/점검자/점검일자 서명란
    장비현황     Hostname / 모델 / IP / S/N / OS / 용도 / 비고 (장비 수만큼 동적으로 팽창)
    점검요약     장비별 특이사항 + 조치 및 점검 소견 (NC 보고서의 '0-1. 요약' 시트 대응)
    <Hostname>  장비 1대당 1시트 — 기본정보 헤더 + 점검항목 표(기준값/전월/당월) + 특이사항

점검 항목의 목록·순서·기준값은 config/inspection_report_template.yaml이 정의하고(코드 수정
없이 양식 변경 가능), 각 항목의 "값을 어떻게 뽑을지"는 이 모듈의 EVALUATORS가 담당한다.

원본로그는 두 가지 포맷을 모두 받는다:
  1) engine/collector.py가 쓴 "--- {cmd} ---" 섹션 포맷
  2) 세션 터미널이 그대로 저장한 프롬프트 트랜스크립트("Core1(config)#show version" 형태)
     — cache/original_log/의 실제 포맷이 이쪽이라, split_transcript()가 프롬프트 줄을
     커맨드 경계로 삼아 잘라낸다.
"""
import re
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.paths import AppPaths
# 판정 상태 문자열은 openpyxl 과 무관하므로 report/inspection_status.py 로 내렸다.
# 여기서 재노출해 기존 import 경로(`from report.inspection_excel import STATUS_OK`)를 지킨다.
from report.inspection_status import (
    STATUS_NA, STATUS_OK, STATUS_UNREACHABLE, STATUS_WARN,
)
from report.textfsm_parser import split_raw_log

__all__ = [
    "load_template", "split_transcript", "evaluate_device", "build_workbook",
    "STATUS_OK", "STATUS_WARN", "STATUS_NA", "STATUS_UNREACHABLE",
]

_TEMPLATE_PATH = "inspection_report_template.yaml"

# 세션 터미널 트랜스크립트의 커맨드 에코 줄: "Core1#show version", "Core1(config)#show mlag".
# 프롬프트 이름은 영문/숫자/점/하이픈/밑줄만 허용해 syslog 줄이 잘못 걸리는 걸 막고,
# (config)·(config-if-Et1) 같은 모드 표시는 선택적으로 흡수한다.
_PROMPT_LINE_RE = re.compile(r"^([\w][\w.\-]{0,60})(?:\([^)\r\n]{0,40}\))?\s*[#>]\s*(.*)$")
# "--- show version ---" 커맨드 헤더. 줄 단위로 매칭하므로 CRLF/LF를 가리지 않는다.
_SECTION_HEADER_RE = re.compile(r"^---\s+(.+?)\s+---\s*$")

_EXCEL_FORBIDDEN_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


# --------------------------------------------------------------------------- 템플릿

def load_template(path=None) -> dict:
    """config/inspection_report_template.yaml을 읽어 dict로 반환.
    파일이 없거나 깨져 있으면 최소 동작 가능한 기본값을 돌려준다 — 보고서 생성이
    설정 파일 하나 때문에 통째로 실패하지 않게 하기 위함."""
    path = Path(path) if path else AppPaths.config_root() / _TEMPLATE_PATH
    try:
        with path.open(encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
    except (OSError, yaml.YAMLError):
        data = {}
    meta = data.get("meta") or {}
    items = data.get("check_items") or []
    if not items:
        items = [{"group": "기본", "name": "System Uptime",
                  "commands": ["show version"], "criteria": "Uptime 확인", "evaluator": "uptime"}]
    # no를 비워둔 항목에 순번을 채운다(템플릿을 손으로 편집할 때 번호를 다시 매기지 않도록).
    for i, item in enumerate(items, start=1):
        item.setdefault("no", i)
        item.setdefault("group", "")
        # 사람이 손으로 고치는 파일이라 'dir flash:'처럼 콜론으로 끝나는 커맨드를 인용부호 없이
        # 쓰면 YAML이 dict으로 읽어버린다 — 그 값이 그대로 셀에 들어가 저장이 깨지므로 여기서
        # 문자열로 정규화한다.
        item["commands"] = [
            next(iter(c)) if isinstance(c, dict) and c else str(c)
            for c in (item.get("commands") or [])
        ]
        item.setdefault("criteria", "")
        item.setdefault("evaluator", "keyword_ok")
    meta.setdefault("title_suffix", "정기점검 보고서")
    meta.setdefault("manager", {})
    meta.setdefault("inspector", {})
    # PDF 보고서(report/inspection_pdf.py)의 표지 전용 필드 — 표준 양식엔 있지만 엑셀
    # 파이프라인엔 없던 개념이라 기본값만 여기서 채우고(코드 수정 없이 yaml로 바꿀 수 있게),
    # 엑셀 렌더러(_build_cover)는 그대로 쓰지 않는다.
    meta.setdefault("confirmer", {})
    meta.setdefault("vendor", "")
    meta.setdefault("report_title", "네트워크 스위치 정기점검")
    meta.setdefault("site_name", "")
    return {"meta": meta, "check_items": items}


# --------------------------------------------------------------------------- 로그 파싱

def split_transcript(raw_text) -> dict:
    """원본로그 한 파일을 {커맨드: 출력} dict로 분해.

    두 가지 저장 포맷을 모두 받는다:
      1) "--- {cmd} ---" 헤더 포맷 (engine/collector.py, 세션 터미널의 커맨드 실행 기록)
      2) 프롬프트 트랜스크립트 ("Core1(config)#show mlag") — 헤더 없이 화면 그대로 저장된 것

    헤더가 하나라도 있으면 헤더를 경계로 삼고, 없을 때만 프롬프트 줄을 경계로 쓴다.
    같은 커맨드가 여러 번 실행됐으면 출력이 있는 마지막 결과를 남긴다 — 페이징 해제 직후처럼
    첫 실행이 빈 출력인 경우 그 빈 값이 실제 출력을 덮어쓰면 안 되기 때문.

    report.textfsm_parser.split_raw_log()를 쓰지 않는 이유: 그 정규식은 헤더 뒤에 "\\n"이
    바로 오는 것을 전제하는데, 실제 cache/original_log/의 파일은 CRLF(\\r\\n)로 저장돼 있어
    한 건도 매칭되지 않는다. 여기서는 splitlines()로 줄을 나눠 개행 방식과 무관하게 처리한다.
    """
    if not raw_text:
        return {}

    lines = raw_text.splitlines()
    result = {}
    current_cmd = None
    buffer = []

    def flush():
        if current_cmd is not None:
            body = "\n".join(buffer).strip("\n")
            if body or current_cmd not in result:
                result[current_cmd] = body

    has_headers = any(_SECTION_HEADER_RE.match(line) for line in lines)

    for line in lines:
        if has_headers:
            match = _SECTION_HEADER_RE.match(line)
            command = match.group(1).strip() if match else None
        else:
            match = _PROMPT_LINE_RE.match(line)
            command = match.group(2).strip() if match else None
            # 프롬프트만 있고 커맨드가 없는 줄과, 구분자로 쓰이는 "!"는 경계가 아니다.
            if match and not command:
                continue
            if command == "!":
                continue

        if command:
            flush()
            current_cmd = command
            buffer = []
            continue
        if match:  # 헤더/프롬프트 줄 자체는 출력에 넣지 않는다.
            continue
        if current_cmd is not None:
            # 헤더 포맷에서는 헤더 바로 다음 줄에 커맨드가 그대로 에코돼 있으므로 건너뛴다.
            if has_headers and not buffer and line.strip() == current_cmd:
                continue
            buffer.append(line)
    flush()
    return result


def _find_section(sections: dict, patterns) -> tuple:
    """patterns(부분 문자열 목록)에 가장 먼저 걸리는 커맨드 섹션을 (커맨드, 출력)으로 반환.
    커맨드 문자열은 장비마다 옵션이 조금씩 달라(`show environment power detail` 등)
    완전 일치가 아니라 부분 일치로 찾는다."""
    lowered = {cmd.lower(): cmd for cmd in sections}
    for pattern in patterns or []:
        needle = str(pattern).lower()
        for low, original in lowered.items():
            if needle in low:
                return original, sections[original]
    return None, None


# --------------------------------------------------------------------------- 평가기(비즈니스 로직)

def _anomaly_lines(text):
    """이상 징후 키워드가 걸린 줄만 반환 — 판정 기준을 원본로그분석 탭과 동일하게 유지하려고
    engine.log_analysis의 키워드/benign 규칙(config/log_rules.json)을 그대로 재사용한다."""
    from engine.log_analysis import classify_line
    hits = []
    for line in (text or "").splitlines():
        keyword = classify_line(line)
        if keyword:
            hits.append((keyword, line.strip()))
    return hits


def _eval_keyword_ok(text, item):
    hits = _anomaly_lines(text)
    if not hits:
        return {"value": STATUS_OK, "status": STATUS_OK}
    preview = "; ".join(f"{kw}: {line}" for kw, line in hits[:3])
    return {"value": f"{STATUS_WARN} ({len(hits)}건)", "status": STATUS_WARN, "detail": preview}


def _eval_uptime(text, item):
    match = re.search(r"[Uu]ptime\s*:?\s*(.+)", text or "")
    if not match:
        # show processes top의 첫 줄 형식: " 10:28:11 up 8 days,  1:23, ..."
        match = re.search(r"\bup\s+([^,]+(?:,\s*[^,]+)?)", text or "")
    if not match:
        return {"value": STATUS_NA, "status": STATUS_NA}
    return {"value": match.group(1).strip().rstrip("."), "status": STATUS_OK}


def _eval_cpu(text, item):
    """show processes top의 CPU 행에서 idle을 찾아 사용률(100-idle)을 계산한다.
    idle이 없으면 us(user) 값이라도 쓴다 — 대한전선 보고서가 '5.4 us'로 기록한 방식."""
    idle = re.search(r"([\d.]+)\s*%?\s*id", text or "")
    if idle:
        used = round(100.0 - float(idle.group(1)), 1)
    else:
        user = re.search(r"([\d.]+)\s*%?\s*us", text or "")
        if not user:
            return {"value": STATUS_NA, "status": STATUS_NA}
        used = round(float(user.group(1)), 1)
    threshold = item.get("warn_at", 70)
    status = STATUS_WARN if used >= float(threshold) else STATUS_OK
    return {"value": f"{used}%", "status": status, "number": used}


def _eval_memory_free(text, item):
    """Free Memory 비율. show processes top의 KiB Mem 행을 먼저 보고, 없으면 show version의
    Total/Free memory(kB)를 쓴다 — LGES 보고서가 Total/Free 바이트로 비율을 산출한 방식."""
    total = free = None
    mem_row = re.search(r"(?:KiB|MiB)\s+Mem\s*:?\s*(.+)", text or "")
    if mem_row:
        row = mem_row.group(1)
        total_m = re.search(r"([\d.]+)\s*total", row)
        free_m = re.search(r"([\d.]+)\s*free", row)
        if total_m and free_m:
            total, free = float(total_m.group(1)), float(free_m.group(1))
    if total is None:
        total_m = re.search(r"Total memory:\s*([\d]+)", text or "")
        free_m = re.search(r"Free memory:\s*([\d]+)", text or "")
        if total_m and free_m:
            total, free = float(total_m.group(1)), float(free_m.group(1))
    if not total:
        return {"value": STATUS_NA, "status": STATUS_NA}
    ratio = round(free / total, 3)
    threshold = float(item.get("warn_at", 0.3))
    status = STATUS_WARN if ratio < threshold else STATUS_OK
    return {"value": ratio, "status": status, "number": ratio, "number_format": "0.0%"}


_ERROR_COUNTER_HEADER_RE = re.compile(r"^\s*Port\s+(.*(?:FCS|CRC|Align|Symbol|Runts|Giants).*)$", re.IGNORECASE)


def _eval_interface_error(text, item):
    """`show interfaces counters errors`의 표를 읽어 0이 아닌 카운터가 있는 포트만 요약한다.
    "Et3/4 FCS 10802"처럼 포트+에러종류+카운트를 붙여 만드는 게 레거시 보고서의 표기 방식."""
    lines = (text or "").splitlines()
    columns = None
    problems = []
    for line in lines:
        header = _ERROR_COUNTER_HEADER_RE.match(line)
        if header:
            columns = header.group(1).split()
            continue
        if not columns:
            continue
        parts = line.split()
        if len(parts) < 2 or not re.match(r"^(Et|Eth|Ma|Po|Vl|Tu)", parts[0]):
            continue
        port, values = parts[0], parts[1:]
        for name, raw in zip(columns, values):
            if not re.fullmatch(r"\d+", raw) or int(raw) == 0:
                continue
            problems.append(f"{port} {name} {raw}")
    if columns is None:
        return {"value": STATUS_NA, "status": STATUS_NA}
    if not problems:
        return {"value": STATUS_OK, "status": STATUS_OK}
    return {"value": ", ".join(problems[:6]) + (" ..." if len(problems) > 6 else ""),
            "status": STATUS_WARN, "detail": "\n".join(problems)}


def _eval_route_summary(text, item):
    """`show ip route summary`의 프로토콜별 라우트 수를 "connected:11 / static:57 / ospf:46"로.
    증감 비교(전월 대비)는 값이 문자열로 그대로 남으므로 보고서에서 눈으로 바로 확인된다."""
    counts = {}
    for proto in ("connected", "static", "ospf", "bgp", "isis", "rip"):
        match = re.search(rf"^\s*{proto}\b[^\d\n]*(\d+)", text or "", re.IGNORECASE | re.MULTILINE)
        if match:
            counts[proto] = int(match.group(1))
    if not counts:
        return {"value": STATUS_NA, "status": STATUS_NA}
    return {"value": "\n".join(f"{k}:{v}" for k, v in counts.items()), "status": STATUS_OK}


def _eval_log_scan(text, item):
    hits = _anomaly_lines(text)
    if not hits:
        return {"value": "없음", "status": STATUS_OK}
    return {"value": f"특이 로그 {len(hits)}건", "status": STATUS_WARN,
            "detail": "\n".join(line for _, line in hits[:20])}


EVALUATORS = {
    "uptime": _eval_uptime,
    "cpu": _eval_cpu,
    "memory_free": _eval_memory_free,
    "interface_error": _eval_interface_error,
    "route_summary": _eval_route_summary,
    "log_scan": _eval_log_scan,
    "keyword_ok": _eval_keyword_ok,
}


def evaluate_device(sections: dict, template: dict) -> list:
    """장비 1대의 {커맨드: 출력}을 템플릿의 점검 항목 순서대로 판정해 행 목록을 만든다.
    반환 각 항목: {no, group, name, method, criteria, value, status, detail}.
    해당 커맨드를 아예 수집하지 않았으면 '미수집'으로 남긴다 — 빈칸으로 두면 정상인지
    안 봤는지 구분이 안 되기 때문."""
    rows = []
    for item in template["check_items"]:
        command, output = _find_section(sections, item.get("commands"))
        if command is None:
            result = {"value": STATUS_NA, "status": STATUS_NA}
        else:
            evaluator = EVALUATORS.get(item.get("evaluator"), _eval_keyword_ok)
            result = evaluator(output, item)
        rows.append({
            "no": item.get("no"), "group": item.get("group", ""), "name": item.get("name", ""),
            "method": command or (item.get("commands") or [""])[0],
            "criteria": item.get("criteria", ""),
            "value": result.get("value"), "status": result.get("status"),
            "detail": result.get("detail", ""),
            "number": result.get("number"), "number_format": result.get("number_format"),
        })
    return rows


# --------------------------------------------------------------------------- 서식

_THIN = Side(style="thin", color="9AA0A6")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FILL_HEAD = PatternFill("solid", start_color="D9E1F2")
FILL_LABEL = PatternFill("solid", start_color="F2F2F2")
FILL_WARN = PatternFill("solid", start_color="FFF2CC")
FILL_BAD = PatternFill("solid", start_color="FCE4E4")
FONT_BAD = Font(bold=True, color="C00000", size=10)
FONT_HEAD = Font(bold=True, size=10)
FONT_BODY = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _put(ws, coord, value, *, font=FONT_BODY, fill=None, align=CENTER, border=BORDER,
          number_format=None):
    """셀 하나에 값 + 서식을 넣는다. 서식을 셀별로 새로 만들지 않고 모듈 상수를 공유하므로
    수백 셀을 채워도 스타일 객체가 늘어나지 않는다."""
    cell = ws[coord]
    cell.value = value
    cell.font = font
    cell.alignment = align
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def _merge(ws, ref, value, **kwargs):
    """병합 셀에 값을 넣는다 — 병합 범위의 왼쪽 위 셀에만 값이 들어가지만, 테두리는
    범위 전체에 그려야 표가 끊겨 보이지 않으므로 모든 셀에 border를 적용한다."""
    ws.merge_cells(ref)
    first = ref.split(":")[0]
    _put(ws, first, value, **kwargs)
    for row in ws[ref]:
        for cell in row:
            cell.border = kwargs.get("border", BORDER)
    return ws[first]


def _widths(ws, mapping):
    for column, width in mapping.items():
        ws.column_dimensions[column].width = width


def _status_style(status):
    """점검 결과 상태에 맞는 (폰트, 배경) — 접속 불가/확인필요는 눈에 바로 띄게."""
    if status == STATUS_UNREACHABLE:
        return FONT_BAD, FILL_BAD
    if status == STATUS_WARN:
        return Font(bold=True, size=10, color="7F6000"), FILL_WARN
    if status == STATUS_NA:
        return Font(size=10, color="808080"), None
    return FONT_BODY, None


def _safe_sheet_title(name, used):
    """엑셀 시트명 제약(31자, : \\ / ? * [ ] 금지)에 맞추고 중복이면 _2, _3을 붙인다."""
    title = _EXCEL_FORBIDDEN_SHEET_CHARS.sub("_", str(name or "장비")).strip() or "장비"
    title = title[:31]
    if title.lower() not in used:
        used.add(title.lower())
        return title
    for suffix in range(2, 100):
        candidate = f"{title[:31 - len(str(suffix)) - 1]}_{suffix}"
        if candidate.lower() not in used:
            used.add(candidate.lower())
            return candidate
    used.add(title.lower())
    return title


# --------------------------------------------------------------------------- 시트 빌더

def _build_cover(ws, ctx):
    """표지 — 제목은 요구사항대로 고객사명 + 프로파일(회차)명으로 조립한다."""
    meta = ctx["template"]["meta"]
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 3, "B": 14, "C": 16, "D": 22, "E": 22, "F": 18, "G": 3})
    for row in range(4, 12):
        ws.row_dimensions[row].height = 34

    title = f"{ctx['customer']}\n{ctx['profile']}\n({meta.get('title_suffix', '정기점검 보고서')})"
    _merge(ws, "B4:F11", title, font=Font(bold=True, size=24), align=CENTER, border=None)

    _merge(ws, "B14:F14", f"점검 대상 {len(ctx['devices'])}대  |  생성 시각 {ctx['generated_at']}",
           font=Font(size=11, color="595959"), align=CENTER, border=None)

    rows = [
        ("담당자", ctx["manager"]),
        ("점검자", ctx["inspector"]),
    ]
    row = 20
    for label, person in rows:
        _merge(ws, f"B{row}:B{row + 1}", label, font=FONT_HEAD, fill=FILL_LABEL)
        for column, header in zip("CDE", ("회사명", "담당자명", "연락처")):
            _put(ws, f"{column}{row}", header, font=FONT_HEAD, fill=FILL_HEAD)
        _put(ws, f"F{row}", "확인", font=FONT_HEAD, fill=FILL_HEAD)
        _put(ws, f"C{row + 1}", person.get("company", ""))
        _put(ws, f"D{row + 1}", person.get("name", ""))
        _put(ws, f"E{row + 1}", person.get("contact", ""))
        _put(ws, f"F{row + 1}", "")
        ws.row_dimensions[row + 1].height = 28
        row += 3

    _put(ws, f"B{row}", "점검일자", font=FONT_HEAD, fill=FILL_LABEL)
    _merge(ws, f"C{row}:F{row}", ctx["inspection_date"], font=Font(bold=True, size=11))
    ws.row_dimensions[row].height = 26


def _build_inventory_sheet(ws, ctx):
    """장비현황 — 장비 수만큼 아래로 동적으로 팽창하는 2차원 매트릭스."""
    headers = ["Hostname", "모델명", "IP", "S/N", "OS Version", "용도/역할", "비고"]
    _widths(ws, {"A": 26, "B": 22, "C": 16, "D": 18, "E": 14, "F": 16, "G": 30})
    _merge(ws, f"A1:{get_column_letter(len(headers))}1", "장비현황",
           font=Font(bold=True, size=13), fill=FILL_HEAD)
    for index, header in enumerate(headers, start=1):
        _put(ws, f"{get_column_letter(index)}2", header, font=FONT_HEAD, fill=FILL_HEAD)

    for offset, device in enumerate(ctx["devices"]):
        row = 3 + offset
        values = [device["name"], device.get("model", ""), device.get("ip", ""),
                  device.get("serial", ""), device.get("os_version", ""),
                  device.get("role", ""), device.get("memo", "")]
        unreachable = device.get("unreachable")
        for index, value in enumerate(values, start=1):
            _put(ws, f"{get_column_letter(index)}{row}", value,
                 font=FONT_BAD if unreachable else FONT_BODY,
                 fill=FILL_BAD if unreachable else None,
                 align=LEFT if index in (1, 7) else CENTER)
        if unreachable:
            ws[f"G{row}"].value = STATUS_UNREACHABLE
    ws.freeze_panes = "A3"


def _build_summary_sheet(ws, ctx):
    """점검요약 — 장비별 특이사항과 소견. NC 보고서의 '요약' 시트가 이 역할을 한다."""
    headers = ["Hostname", "용도/역할", "IP", "점검결과", "특이사항", "조치 및 점검 소견"]
    _widths(ws, {"A": 26, "B": 16, "C": 16, "D": 12, "E": 62, "F": 30})
    _merge(ws, "A1:F1", f"{ctx['customer']} {ctx['profile']} 점검 요약",
           font=Font(bold=True, size=13), fill=FILL_HEAD)
    for index, header in enumerate(headers, start=1):
        _put(ws, f"{get_column_letter(index)}2", header, font=FONT_HEAD, fill=FILL_HEAD)

    for offset, device in enumerate(ctx["devices"]):
        row = 3 + offset
        ws.row_dimensions[row].height = 30
        overall = device.get("overall_status", STATUS_OK)
        font, fill = _status_style(overall)
        _put(ws, f"A{row}", device["name"], align=LEFT)
        _put(ws, f"B{row}", device.get("role", ""))
        _put(ws, f"C{row}", device.get("ip", ""))
        _put(ws, f"D{row}", overall, font=font, fill=fill)
        _put(ws, f"E{row}", device.get("remarks") or "특이사항 없음", align=TOP_LEFT)
        _put(ws, f"F{row}", device.get("opinion", ""), align=TOP_LEFT)
    ws.freeze_panes = "A3"


def _build_device_sheet(ws, ctx, device):
    """장비 1대 상세 시트 — 기본정보(2행) + 점검항목 표 + 특이사항 블록.
    대한전선 보고서의 장비 시트 레이아웃을 그대로 따르되, 항목 수가 템플릿에 따라
    달라지므로 표의 높이는 렌더링 시점에 계산한다."""
    _widths(ws, {"A": 6, "B": 12, "C": 20, "D": 12, "E": 12, "F": 14, "G": 14,
                  "H": 26, "I": 24, "J": 24})

    # ---- 기본정보 (1~2행)
    _merge(ws, "A1:B2", "기본\n정보", font=FONT_HEAD, fill=FILL_HEAD)
    basics = [
        [("고객명", ctx["customer"]), ("점검일시", ctx["inspection_date"]),
         ("장비명/Hostname", device["name"]), ("IP", device.get("ip", ""))],
        [("용도/역할", device.get("role", "")), ("모델명", device.get("model", "")),
         ("S/N", device.get("serial", "")), ("OS Version", device.get("os_version", ""))],
    ]
    for row_index, pairs in enumerate(basics, start=1):
        for column_index, (label, value) in enumerate(pairs):
            label_col = get_column_letter(3 + column_index * 2)
            value_col = get_column_letter(4 + column_index * 2)
            _put(ws, f"{label_col}{row_index}", label, font=FONT_HEAD, fill=FILL_LABEL)
            _put(ws, f"{value_col}{row_index}", value)

    # ---- 점검항목 표 헤더 (3행)
    _put(ws, "A3", "No", font=FONT_HEAD, fill=FILL_HEAD)
    _put(ws, "B3", "구분", font=FONT_HEAD, fill=FILL_HEAD)
    _merge(ws, "C3:D3", "점검항목", font=FONT_HEAD, fill=FILL_HEAD)
    _merge(ws, "E3:G3", "점검 방법", font=FONT_HEAD, fill=FILL_HEAD)
    _put(ws, "H3", "기준값", font=FONT_HEAD, fill=FILL_HEAD)
    _put(ws, "I3", "전월 점검값", font=FONT_HEAD, fill=FILL_HEAD)
    _put(ws, "J3", "당월 점검값", font=FONT_HEAD, fill=FILL_HEAD)

    # ---- 점검항목 행
    first_row = 4
    rows = device.get("items", [])
    for offset, item in enumerate(rows):
        row = first_row + offset
        ws.row_dimensions[row].height = 26
        _put(ws, f"A{row}", item.get("no"))
        _put(ws, f"B{row}", item.get("group", ""))
        _merge(ws, f"C{row}:D{row}", item.get("name", ""), align=LEFT)
        _merge(ws, f"E{row}:G{row}", item.get("method", ""), align=LEFT,
               font=Font(size=9, color="404040"))
        _put(ws, f"H{row}", item.get("criteria", ""), align=LEFT, font=Font(size=9))
        _put(ws, f"I{row}", item.get("previous", ""), align=LEFT)
        font, fill = _status_style(item.get("status"))
        value = item.get("number") if item.get("number") is not None else item.get("value")
        _put(ws, f"J{row}", value, align=LEFT, font=font, fill=fill,
             number_format=item.get("number_format"))

    # ---- '구분' 열의 연속 동일 값 병합 (레거시 보고서의 세로 병합 표현)
    start = first_row
    for offset in range(1, len(rows) + 1):
        row = first_row + offset
        current = rows[offset]["group"] if offset < len(rows) else None
        if current != rows[offset - 1]["group"]:
            if row - 1 > start:
                _merge(ws, f"B{start}:B{row - 1}", rows[offset - 1]["group"],
                       font=FONT_HEAD, fill=FILL_LABEL)
            else:
                _put(ws, f"B{start}", rows[offset - 1]["group"], font=FONT_HEAD, fill=FILL_LABEL)
            start = row

    # ---- 특이사항 블록
    remark_row = first_row + len(rows)
    remark_end = remark_row + 3
    _merge(ws, f"A{remark_row}:B{remark_end}", "특이\n사항", font=FONT_HEAD, fill=FILL_HEAD)
    _merge(ws, f"C{remark_row}:J{remark_end}", device.get("remarks_detail") or "특이사항 없음",
           align=TOP_LEFT, font=Font(size=9))
    for row in range(remark_row, remark_end + 1):
        ws.row_dimensions[row].height = 30

    if device.get("unreachable"):
        _merge(ws, f"C{remark_row}:J{remark_end}",
               f"{STATUS_UNREACHABLE} — 원본로그가 없어 점검값을 수집하지 못했습니다.",
               align=TOP_LEFT, font=FONT_BAD)

    ws.freeze_panes = "A4"
    # 인쇄(=보고서 '페이지') 설정: 장비 1대 = 가로 1페이지. fitToPage는 sheet_properties의
    # pageSetUpPr에 있고 이 객체가 기본 None이라 새로 만들어 넣어야 한다.
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.print_area = f"A1:J{remark_end}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:3"


# --------------------------------------------------------------------------- 엔트리포인트

def build_workbook(ctx: dict) -> Workbook:
    """보고서 컨텍스트(engine.inspection_report_builder.build_context()의 반환값)로
    완성된 Workbook 객체를 만든다. 저장은 호출부 책임(StorageService/파일 경로 정책 분리)."""
    wb = Workbook()
    _build_cover(wb.active, ctx)
    wb.active.title = "표지"
    _build_inventory_sheet(wb.create_sheet("장비현황"), ctx)
    _build_summary_sheet(wb.create_sheet("점검요약"), ctx)

    used_titles = {"표지", "장비현황", "점검요약"}
    for device in ctx["devices"]:
        title = _safe_sheet_title(device["name"], used_titles)
        _build_device_sheet(wb.create_sheet(title), ctx, device)
    return wb
