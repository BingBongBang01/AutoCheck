"""Device Inventory — CSV/YAML/JSON/Excel 불러오기·엑셀 내보내기.
핵심 CRUD/스키마는 device_inventory_core.py 참고.
"""
import csv
import json
import yaml

from engine.device_inventory_core import _normalize_device, update_device


# ---------- Import ----------
def import_csv(path):
    devices = []
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            devices.append(_normalize_device({k.lower().replace(" ", "_"): v for k, v in row.items()}))
    return devices


def import_yaml(path):
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    raw_list = data.get("devices", data) if isinstance(data, dict) else data
    return [_normalize_device(d) for d in raw_list]


def import_json(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    raw_list = data.get("devices", data) if isinstance(data, dict) else data
    return [_normalize_device(d) for d in raw_list]


def import_excel(path):
    """openpyxl 있을 때만 동작. 없으면 안내하고 빈 리스트 반환."""
    try:
        import openpyxl
    except ImportError:
        print("[안내] openpyxl 미설치 — Excel import는 건너뜀 (pip install openpyxl로 추가 가능)")
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).lower().replace(" ", "_") if c.value else "" for c in ws[1]]
    devices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, row))
        raw = {k: v for k, v in raw.items() if k}
        if raw.get("name"):
            devices.append(_normalize_device(raw))
    return devices


def merge_imported(inventory, imported_devices, overwrite=False):
    """import된 장비를 기존 inventory에 합침. overwrite=True면 동명 장비 덮어씀, 아니면 건너뜀."""
    existing_names = {d["name"] for d in inventory["devices"]}
    added, skipped = 0, 0
    for imp in imported_devices:
        if imp["name"] in existing_names:
            if overwrite:
                update_device(inventory, imp["name"], imp)
            else:
                skipped += 1
                continue
        else:
            inventory["devices"].append(imp)
        added += 1
    return {"added": added, "skipped": skipped}


def copy_from_inventory(src_inventory, dst_inventory, overwrite=False):
    """
    다른 정기점검 프로파일의 장비 목록을 현재 프로파일로 가져온다.

    같은 고객사를 반복 점검할 때 장비·IP·계정은 회차가 바뀌어도 대부분 그대로라,
    새 프로파일마다 처음부터 다시 입력하는 게 가장 번거로운 일이었다.
    파일에서 불러오는 것(import_csv 등)과 완전히 같은 병합 규칙을 쓴다 —
    같은 이름의 장비는 overwrite에 따라 덮어쓰거나 건너뛰므로 기본값(False)에서는
    받는 쪽 데이터가 절대 사라지지 않는다.

    받는 쪽에 장비가 하나도 없을 때만 프로젝트 기본값(기본 계정/포트/IP Pool)도 함께
    가져온다 — 이미 장비를 넣어 둔 프로파일의 기본값을 말없이 바꾸면 안 되기 때문.
    """
    was_empty = not dst_inventory.get("devices")
    devices = [_normalize_device(d) for d in src_inventory.get("devices", [])]
    result = merge_imported(dst_inventory, devices, overwrite=overwrite)
    result["defaults_copied"] = False
    if was_empty and src_inventory.get("defaults"):
        dst_inventory["defaults"] = dict(src_inventory["defaults"])
        result["defaults_copied"] = True
    return result


# ---------- Export ----------
EXCEL_HEADERS = ["name", "role", "management_ip", "ssh_port", "username", "password",
                  "vendor", "model", "zone", "site", "location", "warranty",
                  "tag", "memo", "enabled", "auth_method", "key_path"]


def export_to_excel(devices, path):
    """장비목록을 엑셀 한 장으로 내보냄(비밀번호/키 원문 포함 — 로컬 백업/이관용 파일이므로 취급 주의)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl 후 다시 시도하세요.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "device_inventory"
    ws.append(EXCEL_HEADERS)
    for d in devices:
        ws.append([
            d.get("name", ""), d.get("role", ""), d.get("management_ip", ""),
            d.get("ssh_port", 22), d.get("username", ""), d.get("password", ""),
            d.get("vendor", ""), d.get("model", ""), d.get("zone", ""), d.get("site", ""),
            d.get("location", ""), d.get("warranty", ""),
            ",".join(d.get("tag") or []) if isinstance(d.get("tag"), list) else (d.get("tag") or ""),
            d.get("memo", ""), "TRUE" if d.get("enabled") else "FALSE",
            d.get("auth_method", "password"), d.get("key_path", ""),
        ])
    for col, width in zip("ABCDEFGHIJKLMNOPQ",
                          (14, 10, 16, 8, 12, 12, 10, 14, 10, 10, 14, 12, 14, 20, 8, 12, 20)):
        ws.column_dimensions[col].width = width
    wb.save(path)
