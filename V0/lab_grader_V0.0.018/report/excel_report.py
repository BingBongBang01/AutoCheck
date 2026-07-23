"""
Requirement 6 (Excel Export) — 장비별 metrics(dict)를 pandas.DataFrame으로 모은 뒤
.transpose()로 "장비=열, 항목=행" 축으로 뒤집어서 기존 .xlsx 템플릿 좌표에 그대로 써넣는다.
openpyxl로 셀 값만 갱신하므로 템플릿에 미리 잡혀있는 스타일(폰트/테두리/색)은 그대로 유지된다.
"""
import os
import openpyxl
import pandas as pd

UNREACHABLE_LABEL = "접속 불가"

# 보고서에 노출할 항목(행) 순서 및 라벨 — 템플릿 쪽 좌측 라벨과 맞춰서 사용.
ATTRIBUTE_LABELS = [
    ("model", "모델명"),
    ("serial_number", "시리얼 번호"),
    ("image", "소프트웨어 버전"),
    ("uptime", "가동 시간(Uptime)"),
    ("cpu_util_percent", "CPU 사용률(%)"),
    ("memory_used_percent", "메모리 사용률(%)"),
    ("load_average_1m", "Load Average(1분)"),
]


def build_dataframe(dataset):
    """
    dataset: {device_name: metrics_dict} (report.textfsm_parser.build_report_dataset 결과)
    반환: index=항목 라벨, columns=장비명 인 DataFrame (이미 transpose된 상태).
    """
    rows = {}
    for device, metrics in dataset.items():
        if metrics.get("unreachable"):
            rows[device] = {label: UNREACHABLE_LABEL for _, label in ATTRIBUTE_LABELS}
            continue
        rows[device] = {label: metrics.get(key, "") for key, label in ATTRIBUTE_LABELS}
    # {device: {label: value}} -> DataFrame(index=device, columns=label) -> .T로 축 반전(요구사항: 장비=열)
    df = pd.DataFrame.from_dict(rows, orient="index")
    return df.T


def write_into_template(dataset, template_path, output_path, sheet_name=None, start_row=2, start_col=2):
    """
    template_path의 기존 서식(폰트/테두리/색)을 유지한 채, start_row/start_col 좌표부터
    DataFrame(장비=열, 항목=행)을 셀 값으로만 채워 넣는다.
    - 1열(= start_col - 1)에는 항목 라벨, 1행(= start_row - 1)에는 장비명을 쓴다.
    - template이 없으면 새 워크북을 만들어 같은 좌표 규칙으로 채운다(최초 실행용).
    """
    df = build_dataframe(dataset)

    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    ws.cell(row=start_row - 1, column=start_col - 1, value="항목 / 장비")
    for col_idx, device in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row - 1, column=col_idx, value=device)

    for row_offset, label in enumerate(df.index):
        r = start_row + row_offset
        ws.cell(row=r, column=start_col - 1, value=label)
        for col_idx, device in enumerate(df.columns, start=start_col):
            value = df.at[label, device]
            ws.cell(row=r, column=col_idx, value=value)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    sample_dataset = {
        "Core1": {"model": "DCS-7050SX3-48YC8", "serial_number": "SN1", "image": "4.28.0F",
                  "uptime": "3 weeks", "cpu_util_percent": 16.0, "memory_used_percent": 62.5,
                  "load_average_1m": "0.10"},
        "Core2": {"unreachable": True},
    }
    out = write_into_template(sample_dataset, template_path=None, output_path="AutoCheck_report_sample.xlsx")
    print("작성됨:", out)
