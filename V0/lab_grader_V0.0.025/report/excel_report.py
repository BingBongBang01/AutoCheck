"""
Requirement 6 (Excel Export) — 장비별 metrics(dict)를 pandas.DataFrame으로 모은 뒤
.transpose()로 "장비=열, 항목=행" 축으로 뒤집어서 기존 .xlsx 템플릿 좌표에 그대로 써넣는다.
openpyxl로 셀 값만 갱신하므로 템플릿에 미리 잡혀있는 스타일(폰트/테두리/색)은 그대로 유지된다.
"""
import os
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

UNREACHABLE_LABEL = "접속 불가"

# --- 요구사항 (전체 Excel 리포트: Dashboard/Node Inventory/Findings 3-sheet) 공통 스타일 ---
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
RESULT_FILL = {
    "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "UNKNOWN": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}


def _write_header_row(ws, headers, row=1):
    """헤더 행 하나를 굵게+어두운 배경으로 쓰고, autofilter/freeze까지 함께 설정."""
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def _autofit_columns(ws):
    """각 열의 최대 셀 길이 기준으로 대략적인 너비 자동 조정(openpyxl엔 진짜 auto-width가 없어 근사치)."""
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = min(width + 2, 60)


def _fill_result_cell(cell):
    """PASS=초록/FAIL=빨강/UNKNOWN=노랑 — Result 값에 맞는 배경색을 셀에 적용."""
    fill = RESULT_FILL.get(str(cell.value).upper())
    if fill:
        cell.fill = fill

# 보고서에 노출할 항목(행) 순서 및 라벨 — 템플릿 쪽 좌측 라벨과 맞춰서 사용.
ATTRIBUTE_LABELS = [
    ("model", "모델명"),
    ("serial_number", "시리얼 번호"),
    ("image", "소프트웨어 버전"),
    ("uptime", "가동 시간(Uptime)"),
    ("cpu_util_percent", "CPU 사용률(%)"),
    ("memory_used_percent", "메모리 사용률(%)"),
    ("load_average_1m", "Load Average(1분)"),
]


def build_dataframe(dataset):
    """
    dataset: {device_name: metrics_dict} (report.textfsm_parser.build_report_dataset 결과)
    반환: index=항목 라벨, columns=장비명 인 DataFrame (이미 transpose된 상태).
    """
    rows = {}
    for device, metrics in dataset.items():
        if metrics.get("unreachable"):
            rows[device] = {label: UNREACHABLE_LABEL for _, label in ATTRIBUTE_LABELS}
            continue
        rows[device] = {label: metrics.get(key, "") for key, label in ATTRIBUTE_LABELS}
    # {device: {label: value}} -> DataFrame(index=device, columns=label) -> .T로 축 반전(요구사항: 장비=열)
    df = pd.DataFrame.from_dict(rows, orient="index")
    return df.T


def write_into_template(dataset, template_path, output_path, sheet_name=None, start_row=2, start_col=2):
    """
    template_path의 기존 서식(폰트/테두리/색)을 유지한 채, start_row/start_col 좌표부터
    DataFrame(장비=열, 항목=행)을 셀 값으로만 채워 넣는다.
    - 1열(= start_col - 1)에는 항목 라벨, 1행(= start_row - 1)에는 장비명을 쓴다.
    - template이 없으면 새 워크북을 만들어 같은 좌표 규칙으로 채운다(최초 실행용).
    """
    df = build_dataframe(dataset)

    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    ws.cell(row=start_row - 1, column=start_col - 1, value="항목 / 장비")
    for col_idx, device in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row - 1, column=col_idx, value=device)

    for row_offset, label in enumerate(df.index):
        r = start_row + row_offset
        ws.cell(row=r, column=start_col - 1, value=label)
        for col_idx, device in enumerate(df.columns, start=start_col):
            value = df.at[label, device]
            ws.cell(row=r, column=col_idx, value=value)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def build_full_report_workbook(data, reachability=None):
    """
    api.report_api.ReportApiMixin.export_report_excel()이 반환하는 data dict를 받아
    Dashboard / Node Inventory / Findings 3개 시트로 구성된 openpyxl Workbook을 만들어 반환한다.
    저장(save)은 호출자 책임 — 여기선 워크북 객체만 만든다(저장 다이얼로그는 별도 구현 예정).

    reachability: {device_name: bool} — 선택. 없으면 Node Inventory의 Reachability 칸은 "-".
    """
    findings = data.get("findings", [])
    devices = data.get("devices", [])
    health = data.get("health") or {"project_score": None, "device_scores": {}}
    device_scores = health.get("device_scores", {})
    rule_based = data.get("rule_based") or {}

    result_counts_by_device = {}
    for f in findings:
        device = f.get("device", "-")
        counts = result_counts_by_device.setdefault(device, {"PASS": 0, "FAIL": 0, "UNKNOWN": 0})
        result = f.get("result")
        if result in counts:
            counts[result] += 1

    wb = openpyxl.Workbook()

    # --- Sheet 1: Dashboard ---
    ws1 = wb.active
    ws1.title = "Dashboard"
    critical_count = sum(1 for f in findings if f.get("result") == "FAIL")
    warning_count = sum(1 for f in findings if f.get("result") == "UNKNOWN")
    dashboard_rows = [
        ("Project Name", data.get("project_id", "")),
        ("Inspection Time", data.get("session", "")),
        ("Total Nodes", len(devices)),
        ("Overall Health Score", health.get("project_score")),
        ("Critical Count", critical_count),
        ("Warning Count", warning_count),
        ("AI Summary", data.get("ai_summary", "")),
    ]
    for row_idx, (label, value) in enumerate(dashboard_rows, start=1):
        cell = ws1.cell(row=row_idx, column=1, value=label)
        cell.font = Font(bold=True)
        ws1.cell(row=row_idx, column=2, value=value)

    top_priority_row = len(dashboard_rows) + 2
    ws1.cell(row=top_priority_row, column=1, value="Top 5 Priority Actions").font = Font(bold=True)
    priority_headers = ["Priority", "Device", "Check ID", "Suggested Action"]
    _write_header_row(ws1, priority_headers, row=top_priority_row + 1)
    for offset, item in enumerate(rule_based.get("top_priority", [])[:5], start=1):
        r = top_priority_row + 1 + offset
        ws1.cell(row=r, column=1, value=item.get("priority"))
        ws1.cell(row=r, column=2, value=item.get("device"))
        ws1.cell(row=r, column=3, value=item.get("check"))
        ws1.cell(row=r, column=4, value=item.get("suggested_action"))
    _autofit_columns(ws1)

    # --- Sheet 2: Node Inventory ---
    ws2 = wb.create_sheet("Node Inventory")
    node_headers = ["Device Name", "IP Address", "Vendor", "Reachability", "Health Score", "PASS", "FAIL", "UNKNOWN"]
    _write_header_row(ws2, node_headers)
    for row_idx, device in enumerate(devices, start=2):
        name = device.get("name", "")
        counts = result_counts_by_device.get(name, {"PASS": 0, "FAIL": 0, "UNKNOWN": 0})
        reach = reachability.get(name) if reachability is not None else None
        reach_label = "-" if reach is None else ("Reachable" if reach else "Unreachable")
        ws2.cell(row=row_idx, column=1, value=name)
        ws2.cell(row=row_idx, column=2, value=device.get("management_ip", ""))
        ws2.cell(row=row_idx, column=3, value=device.get("vendor", ""))
        ws2.cell(row=row_idx, column=4, value=reach_label)
        ws2.cell(row=row_idx, column=5, value=device_scores.get(name))
        ws2.cell(row=row_idx, column=6, value=counts["PASS"])
        ws2.cell(row=row_idx, column=7, value=counts["FAIL"])
        ws2.cell(row=row_idx, column=8, value=counts["UNKNOWN"])
    _autofit_columns(ws2)

    # --- Sheet 3: Findings ---
    ws3 = wb.create_sheet("Findings")
    finding_headers = ["Node", "Stage", "Check ID", "Result", "Expected", "Actual", "Evidence", "AI Suggested Action"]
    _write_header_row(ws3, finding_headers)
    for row_idx, f in enumerate(findings, start=2):
        ws3.cell(row=row_idx, column=1, value=f.get("device"))
        ws3.cell(row=row_idx, column=2, value=f.get("category"))
        ws3.cell(row=row_idx, column=3, value=f.get("check_id"))
        result_cell = ws3.cell(row=row_idx, column=4, value=f.get("result"))
        ws3.cell(row=row_idx, column=5, value=f.get("expected"))
        ws3.cell(row=row_idx, column=6, value=f.get("actual"))
        ws3.cell(row=row_idx, column=7, value=f.get("evidence"))
        ws3.cell(row=row_idx, column=8, value=f.get("suggested_action"))
        _fill_result_cell(result_cell)
    _autofit_columns(ws3)

    return wb


if __name__ == "__main__":
    sample_dataset = {
        "Core1": {"model": "DCS-7050SX3-48YC8", "serial_number": "SN1", "image": "4.28.0F",
                  "uptime": "3 weeks", "cpu_util_percent": 16.0, "memory_used_percent": 62.5,
                  "load_average_1m": "0.10"},
        "Core2": {"unreachable": True},
    }
    out = write_into_template(sample_dataset, template_path=None, output_path="AutoCheck_report_sample.xlsx")
    print("작성됨:", out)
