"""Device Inventory — CSV/YAML/JSON/Excel 불러오기·엑셀 내보내기.
핵심 CRUD/스키마는 device_inventory_core.py 참고.
"""
import csv
import json
import yaml

from engine.device_inventory_core import _normalize_device, update_device


# ---------- Import ----------
def import_csv(path):
    devices = []
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            devices.append(_normalize_device({k.lower().replace(" ", "_"): v for k, v in row.items()}))
    return devices


def import_yaml(path):
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    raw_list = data.get("devices", data) if isinstance(data, dict) else data
    return [_normalize_device(d) for d in raw_list]


def import_json(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    raw_list = data.get("devices", data) if isinstance(data, dict) else data
    return [_normalize_device(d) for d in raw_list]


def import_excel(path):
    """openpyxl 있을 때만 동작. 없으면 안내하고 빈 리스트 반환."""
    try:
        import openpyxl
    except ImportError:
        print("[안내] openpyxl 미설치 — Excel import는 건너뜀 (pip install openpyxl로 추가 가능)")
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).lower().replace(" ", "_") if c.value else "" for c in ws[1]]
    devices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, row))
        raw = {k: v for k, v in raw.items() if k}
        if raw.get("name"):
            devices.append(_normalize_device(raw))
    return devices


def merge_imported(inventory, imported_devices, overwrite=False):
    """import된 장비를 기존 inventory에 합침. overwrite=True면 동명 장비 덮어씀, 아니면 건너뜀."""
    existing_names = {d["name"] for d in inventory["devices"]}
    added, skipped = 0, 0
    for imp in imported_devices:
        if imp["name"] in existing_names:
            if overwrite:
                update_device(inventory, imp["name"], imp)
            else:
                skipped += 1
                continue
        else:
            inventory["devices"].append(imp)
        added += 1
    return {"added": added, "skipped": skipped}


# ---------- Export ----------
EXCEL_HEADERS = ["name", "role", "management_ip", "ssh_port", "username", "password",
                  "vendor", "model", "zone", "site", "tag", "memo", "enabled",
                  "auth_method", "key_path"]


def export_to_excel(devices, path):
    """장비목록을 엑셀 한 장으로 내보냄(비밀번호/키 원문 포함 — 로컬 백업/이관용 파일이므로 취급 주의)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl 후 다시 시도하세요.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "device_inventory"
    ws.append(EXCEL_HEADERS)
    for d in devices:
        ws.append([
            d.get("name", ""), d.get("role", ""), d.get("management_ip", ""),
            d.get("ssh_port", 22), d.get("username", ""), d.get("password", ""),
            d.get("vendor", ""), d.get("model", ""), d.get("zone", ""), d.get("site", ""),
            ",".join(d.get("tag") or []) if isinstance(d.get("tag"), list) else (d.get("tag") or ""),
            d.get("memo", ""), "TRUE" if d.get("enabled") else "FALSE",
            d.get("auth_method", "password"), d.get("key_path", ""),
        ])
    for col, width in zip("ABCDEFGHIJKLMNO", (14, 10, 16, 8, 12, 12, 10, 14, 10, 10, 14, 20, 8, 12, 20)):
        ws.column_dimensions[col].width = width
    wb.save(path)
