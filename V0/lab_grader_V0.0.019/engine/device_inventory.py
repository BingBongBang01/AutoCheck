"""
Device Inventory — 프로젝트의 장비 정보(IP·계정·역할·벤더 등)를 관리하는 유일한 출처.

기존엔 lab_meta.yaml(name/role)과 ip_allocation.yaml(ip/계정)로 쪼개져 있었고,
Collection이 그 두 파일을 직접 읽어 SSH 접속했음 — 이게 "IP가 코드 여기저기 흩어짐" 문제의 원인.
이제 device_inventory.yaml 하나로 통합하고, Collection은 이 모듈이 제공하는
get_enabled_devices()만 사용해서 접속 대상을 얻는다 (IP를 직접 다루지 않음).

기존 프로젝트(lab_meta.yaml + ip_allocation.yaml만 있던 것)는 최초 load 시 자동 마이그레이션.
"""
import os
import re
import csv
import json
import yaml

DEVICE_FIELDS = ["name", "role", "management_ip", "ssh_port", "username", "password",
                  "vendor", "model", "zone", "site", "tag", "memo", "enabled",
                  "auth_method", "key_path", "key_content", "key_passphrase"]

DEFAULT_DEVICE = {
    "name": "", "role": "", "management_ip": "", "ssh_port": 22,
    "username": "", "password": "", "vendor": "Arista", "model": "",
    "zone": "", "site": "", "tag": [], "memo": "", "enabled": True,
    "auth_method": "password", "key_path": "", "key_content": "", "key_passphrase": "",
}

DEFAULT_PROJECT_DEFAULTS = {
    "management_network": "",
    "default_ssh_port": 22,
    "default_username": "admin",
    "default_password": "admin",
    "ip_pool": {"prefix": "", "start": 0, "end": 0},
}


def _normalize_device(raw):
    d = dict(DEFAULT_DEVICE)
    d.update({k: v for k, v in raw.items() if k in DEVICE_FIELDS})
    return d


def _migrate_from_legacy(lab_meta_path, ip_allocation_path):
    """구버전(lab_meta.yaml + ip_allocation.yaml)에서 device_inventory 구조로 변환."""
    devices = []
    lab_meta, ip_alloc = {}, {}
    if os.path.exists(lab_meta_path):
        with open(lab_meta_path, encoding="utf-8") as f:
            lab_meta = yaml.safe_load(f) or {}
    if os.path.exists(ip_allocation_path):
        with open(ip_allocation_path, encoding="utf-8") as f:
            ip_alloc = yaml.safe_load(f) or {}

    alloc_by_name = {a.get("node_name"): a for a in ip_alloc.get("allocations", [])}
    for d in lab_meta.get("devices", []):
        alloc = alloc_by_name.get(d["name"], {})
        devices.append(_normalize_device({
            "name": d["name"], "role": d.get("role", ""),
            "management_ip": alloc.get("ip", ""),
            "username": alloc.get("username", "") or "",
            "password": alloc.get("password", "") or "",
            "vendor": "Arista", "model": d.get("platform", ""),
            "enabled": True,
        }))

    defaults = dict(DEFAULT_PROJECT_DEFAULTS)
    creds = ip_alloc.get("default_credentials", {})
    if creds:
        defaults["default_username"] = creds.get("username", "admin")
        defaults["default_password"] = creds.get("password", "admin")

    return {"defaults": defaults, "devices": devices}


def load_inventory(inventory_path, lab_meta_path=None, ip_allocation_path=None):
    """
    device_inventory.yaml이 있으면 그대로 로드.
    없고 구버전 파일이 있으면 자동 마이그레이션(파일도 새로 저장해서 다음부턴 바로 로드).
    """
    if os.path.exists(inventory_path):
        with open(inventory_path, encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        data.setdefault("defaults", dict(DEFAULT_PROJECT_DEFAULTS))
        data["devices"] = [_normalize_device(d) for d in data.get("devices", [])]
        return data

    if lab_meta_path and ip_allocation_path:
        migrated = _migrate_from_legacy(lab_meta_path, ip_allocation_path)
        save_inventory(migrated, inventory_path)
        return migrated

    return {"defaults": dict(DEFAULT_PROJECT_DEFAULTS), "devices": []}


def save_inventory(inventory, inventory_path):
    os.makedirs(os.path.dirname(inventory_path), exist_ok=True)
    with open(inventory_path, "w", encoding="utf-8") as f:
        yaml.dump(inventory, f, allow_unicode=True, sort_keys=False)


def resolve_credentials(device, defaults):
    """장비별 값이 비어있으면 프로젝트 기본값을 적용."""
    ip = device.get("management_ip") or ""
    port = device.get("ssh_port") or defaults.get("default_ssh_port", 22)
    username = device.get("username") or defaults.get("default_username", "admin")
    password = device.get("password") or defaults.get("default_password", "admin")
    return ip, port, username, password


def get_enabled_devices(inventory):
    """Collection이 사용할 접속 대상 — enabled=True인 것만."""
    return [d for d in inventory["devices"] if d.get("enabled", True)]


def add_device(inventory, device_dict):
    inventory["devices"].append(_normalize_device(device_dict))


def remove_device(inventory, name):
    inventory["devices"] = [d for d in inventory["devices"] if d["name"] != name]


def update_device(inventory, name, fields):
    for d in inventory["devices"]:
        if d["name"] == name:
            d.update({k: v for k, v in fields.items() if k in DEVICE_FIELDS})
            return True
    return False


# ---------- IP Pool 자동 할당 ----------
def generate_ip_pool(prefix, start, end):
    return [f"{prefix}{i}" for i in range(start, end + 1)]


def auto_allocate_ips(inventory, prefix=None, start=None, end=None):
    """
    management_ip가 비어있는 장비들에게 순서대로 IP Pool을 할당.
    인자를 안 주면 inventory["defaults"]["ip_pool"] 사용.
    반환: 실제로 할당된 개수.
    """
    pool_cfg = inventory.get("defaults", {}).get("ip_pool", {})
    prefix = prefix or pool_cfg.get("prefix", "")
    start = start if start is not None else pool_cfg.get("start", 0)
    end = end if end is not None else pool_cfg.get("end", 0)

    pool = generate_ip_pool(prefix, start, end)
    used_ips = {d["management_ip"] for d in inventory["devices"] if d.get("management_ip")}
    available = [ip for ip in pool if ip not in used_ips]

    allocated = 0
    for d in inventory["devices"]:
        if not d.get("management_ip") and available:
            d["management_ip"] = available.pop(0)
            allocated += 1

    # 빈 IP 행을 다 채우고도 남은 IP는 새 장비 행으로 생성한다.
    # 기존에는 여기서 남은 IP를 버렸기 때문에, 장비가 없거나 부족한 상태에서
    # 자동 생성을 누르면 테이블에 아무것도 채워지지 않는 것처럼 보였다.
    existing_names = {d["name"] for d in inventory["devices"]}
    for ip in available:
        suffix = ip.rsplit(".", 1)[-1] if "." in ip else ip
        name = f"AUTO-{suffix}"
        while name in existing_names:
            suffix += "_"
            name = f"AUTO-{suffix}"
        existing_names.add(name)
        add_device(inventory, {"name": name, "management_ip": ip, "enabled": True})
        allocated += 1

    return allocated


# ---------- Import ----------
def import_csv(path):
    devices = []
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            devices.append(_normalize_device({k.lower().replace(" ", "_"): v for k, v in row.items()}))
    return devices


def import_yaml(path):
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    raw_list = data.get("devices", data) if isinstance(data, dict) else data
    return [_normalize_device(d) for d in raw_list]


def import_json(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    raw_list = data.get("devices", data) if isinstance(data, dict) else data
    return [_normalize_device(d) for d in raw_list]


def import_excel(path):
    """openpyxl 있을 때만 동작. 없으면 안내하고 빈 리스트 반환."""
    try:
        import openpyxl
    except ImportError:
        print("[안내] openpyxl 미설치 — Excel import는 건너뜀 (pip install openpyxl로 추가 가능)")
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).lower().replace(" ", "_") if c.value else "" for c in ws[1]]
    devices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, row))
        raw = {k: v for k, v in raw.items() if k}
        if raw.get("name"):
            devices.append(_normalize_device(raw))
    return devices


def merge_imported(inventory, imported_devices, overwrite=False):
    """import된 장비를 기존 inventory에 합침. overwrite=True면 동명 장비 덮어씀, 아니면 건너뜀."""
    existing_names = {d["name"] for d in inventory["devices"]}
    added, skipped = 0, 0
    for imp in imported_devices:
        if imp["name"] in existing_names:
            if overwrite:
                update_device(inventory, imp["name"], imp)
            else:
                skipped += 1
                continue
        else:
            inventory["devices"].append(imp)
        added += 1
    return {"added": added, "skipped": skipped}


# ---------- 도달가능성 체크 (Dashboard의 Reachable/Offline용) ----------
import socket as _socket


def _check_one(ip, port, timeout):
    if not ip:
        return False
    try:
        with _socket.create_connection((ip, port), timeout=timeout):
            return True
    except Exception:
        return False


def check_reachability(devices, defaults, timeout=2):
    """socket 레벨 포트 체크만 — SSH 인증은 안 함(가벼움). {"name": bool} 반환."""
    result = {}
    for d in devices:
        ip, port, _, _ = resolve_credentials(d, defaults)
        result[d["name"]] = _check_one(ip, port, timeout)
    return result
