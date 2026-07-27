"""
검사 커맨드 카탈로그 관리.
- 필수(essential): 기본 활성화(enabled=True), 어떤 장비 점검이든 공통으로 쓰는 것
- 선택(optional): 기본 비활성화(enabled=False), 상황에 따라 켜서 쓰는 것
- custom: 사용자가 직접 추가한 것

로컬 파일(commands_catalog.yaml)에 저장/불러오기 — 토글·추가·삭제한 상태가 다음 실행 때도 유지됨.
"""
import os
import yaml

DEFAULT_CATALOG_PATH = "config/commands_catalog.yaml"

CATEGORY_LABELS = {"essential": "필수", "optional": "선택사항", "custom": "커스텀"}
EXCEL_HEADERS = ["순서", "카테고리", "사용여부", "커맨드", "설명", "체크ID"]

# 최초 실행 시(파일이 없을 때) 채워 넣을 기본값.
# 근거: 실제 회사 정기점검 보고서 4종 + EOS 매뉴얼 대조 분석 결과.
# check_id: 벤더 무관 추상 ID(plugins/vendors/arista.py의 COMMAND_MAP과 1:1 매칭).
#           command 필드는 하위호환용으로 남겨두되, check_id가 있으면 실행 시 그걸 우선 사용.
DEFAULT_ESSENTIAL = [
    {"check_id": "version_info", "command": "show version", "description": "가동시간·모델 확인", "enabled": True},
    {"check_id": "power_status", "command": "show environment power", "description": "전원 이중화 확인", "enabled": True},
    {"check_id": "cooling_status", "command": "show environment cooling", "description": "팬 상태 확인", "enabled": True},
    {"check_id": "temperature_status", "command": "show environment temperature", "description": "온도 이상 확인", "enabled": True},
    {"check_id": "cpu_usage", "command": "show processes top once", "description": "CPU/메모리 사용률", "enabled": True},
    {"check_id": "log_check", "command": "show log", "description": "특이 로그 확인", "enabled": True},
    {"check_id": "interface_status", "command": "show interface status", "description": "포트 링크 상태", "enabled": True},
    {"check_id": "interface_errors", "command": "show interfaces counters errors", "description": "CRC/에러 카운터", "enabled": True},
    {"check_id": "interface_transceiver", "command": "show interfaces transceiver", "description": "광모듈 정상범위", "enabled": True},
    {"check_id": "interface_brief", "command": "show ip interface brief", "description": "인터페이스 요약", "enabled": True},
    {"check_id": "clock_status", "command": "show clock", "description": "시간 동기화 확인", "enabled": True},
]

DEFAULT_OPTIONAL = [
    {"check_id": "module_status", "command": "show module", "description": "섀시형 장비 모듈 확인", "enabled": False},
    {"check_id": "port_channel_status", "command": "show port-channel summary", "description": "LACP 이중화 확인", "enabled": False},
    {"check_id": "mlag_status", "command": "show mlag", "description": "MLAG 페어 상태", "enabled": False},
    {"check_id": "vrrp_status", "command": "show vrrp brief", "description": "VRRP 이중화 확인", "enabled": False},
    {"check_id": "varp_status", "command": "show ip virtual-router", "description": "VARP 이중화 확인", "enabled": False},
    {"check_id": "stp_status", "command": "show spanning-tree vlan 1,100,200,999", "description": "STP 루트 확인", "enabled": False},
    {"check_id": "evpn_summary", "command": "show bgp evpn summary", "description": "EVPN 네이버 확인", "enabled": False},
    {"check_id": "bgp_summary", "command": "show ip bgp summary", "description": "BGP 네이버 확인", "enabled": False},
    {"check_id": "ospf_neighbor", "command": "show ip ospf neighbor", "description": "OSPF 네이버 확인", "enabled": False},
    {"check_id": "acl_status", "command": "show ip access-lists", "description": "ACL 설정/명시적 deny 확인", "enabled": False},
    {"check_id": "arp_status", "command": "show ip arp vrf all", "description": "ARP 상태 수집", "enabled": False},
    {"check_id": "inventory_status", "command": "show inventory", "description": "S/N·부품 확인", "enabled": False},
    {"check_id": "running_config", "command": "show running-config", "description": "설정 변경 이력 대조용", "enabled": False},
    {"check_id": "reload_cause", "command": "show reload cause", "description": "마지막 재부팅 원인", "enabled": False},
    {"check_id": "ntp_status", "command": "show ntp status", "description": "NTP 동기화 확인", "enabled": False},
    {"check_id": "interface_rates", "command": "show interfaces counters rates", "description": "실시간 트래픽량", "enabled": False},
    {"check_id": "interface_description", "command": "show interfaces description", "description": "포트 라벨 확인", "enabled": False},
]


def _make_default_catalog():
    catalog = []
    for i, item in enumerate(DEFAULT_ESSENTIAL):
        catalog.append({"id": f"essential_{i}", "category": "essential", **item})
    for i, item in enumerate(DEFAULT_OPTIONAL):
        catalog.append({"id": f"optional_{i}", "category": "optional", **item})
    return catalog


def load_catalog(path=DEFAULT_CATALOG_PATH):
    if not os.path.exists(path):
        catalog = _make_default_catalog()
        save_catalog(catalog, path)
        return catalog
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f) or []


def save_catalog(catalog, path=DEFAULT_CATALOG_PATH):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        yaml.dump(catalog, f, allow_unicode=True, sort_keys=False)


def toggle_command(catalog, command_id, enabled):
    for item in catalog:
        if item["id"] == command_id:
            item["enabled"] = enabled
            return True
    return False


def add_command(catalog, command_text, description="", check_id=None):
    existing_ids = [item["id"] for item in catalog if item["category"] == "custom"]
    next_num = len(existing_ids)
    new_id = f"custom_{next_num}"
    while new_id in existing_ids:
        next_num += 1
        new_id = f"custom_{next_num}"
    catalog.append({
        "id": new_id, "category": "custom", "check_id": check_id,
        "command": command_text, "description": description or "(설명 없음)",
        "enabled": True,
    })
    return new_id


def move_items(catalog, item_ids, target_index=None):
    """item_ids(여러 개 가능, 상대 순서 유지)를 카탈로그 전체(단일 목록) 기준
    target_index 위치로 이동. target_index=None이면 맨 끝. catalog는 in-place로 수정됨.
    카탈로그 파일을 매번 새로 읽어 처리하므로(_load_catalog) 클라이언트가 들고 있던
    stale한 DOM 상태와 무관하게 항상 디스크의 최신 목록을 기준으로 동작한다."""
    id_set = set(item_ids)
    moving = [item for item in catalog if item["id"] in id_set]
    if not moving:
        return
    remaining = [item for item in catalog if item["id"] not in id_set]
    insert_pos = len(remaining) if target_index is None else max(0, min(target_index, len(remaining)))
    remaining[insert_pos:insert_pos] = moving
    catalog[:] = remaining


def set_category(catalog, item_id, category):
    """카테고리(필수/선택/커스텀) 라벨만 바꾼다 — 통합 목록에서의 위치는 그대로 유지."""
    for item in catalog:
        if item["id"] == item_id:
            item["category"] = category
            return True
    return False


def remove_command(catalog, command_id):
    idx = next((i for i, item in enumerate(catalog) if item["id"] == command_id), None)
    if idx is None:
        return False
    del catalog[idx]
    return True


def get_enabled_commands(catalog):
    """하위호환용 — literal command 필드 그대로 반환(벤더 무관, 기존 호출부 안 건드림)."""
    return [item["command"] for item in catalog if item["enabled"]]


def get_enabled_commands_via_driver(catalog, vendor_driver):
    """
    신규 — check_id가 있으면 VendorDriver로 실제 커맨드를 resolve.
    check_id가 없는 항목(사용자가 직접 입력한 커스텀 커맨드 등)은 literal command로 폴백.
    """
    commands = []
    for item in catalog:
        if not item["enabled"]:
            continue
        check_id = item.get("check_id")
        resolved = vendor_driver.command_for(check_id) if check_id else None
        commands.append(resolved or item["command"])
    return commands


def export_to_excel(catalog, path):
    """카탈로그를 엑셀 한 장으로 내보냄. 열: 순서, 카테고리, 사용여부, 커맨드, 설명, 체크ID."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl 후 다시 시도하세요.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "commands_catalog"
    ws.append(EXCEL_HEADERS)
    for i, item in enumerate(catalog, start=1):
        ws.append([
            i,
            CATEGORY_LABELS.get(item.get("category"), item.get("category")),
            "TRUE" if item.get("enabled") else "FALSE",
            item.get("command", ""),
            item.get("description", ""),
            item.get("check_id") or "",
        ])
    for col, width in zip("ABCDEF", (6, 12, 10, 32, 30, 20)):
        ws.column_dimensions[col].width = width
    wb.save(path)


def import_from_excel(path):
    """엑셀(export_to_excel 형식)을 읽어 새 catalog 리스트로 반환 — 기존 카탈로그를 통째로 대체할 때 씀."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl 후 다시 시도하세요.")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    label_to_category = {v: k for k, v in CATEGORY_LABELS.items()}
    truthy = {"true", "1", "y", "yes", "o", "사용", "예", "checked"}

    def parse_enabled(v):
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in truthy

    parsed = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[3]:
            continue
        order, category_label, enabled_raw, command, description, check_id = (list(row) + [None] * 6)[:6]
        category = label_to_category.get(str(category_label).strip(), "custom") if category_label else "custom"
        parsed.append({
            "order": order if isinstance(order, (int, float)) else len(parsed) + 1,
            "category": category,
            "enabled": parse_enabled(enabled_raw),
            "command": str(command).strip(),
            "description": str(description).strip() if description else "",
            "check_id": str(check_id).strip() if check_id else None,
        })
    parsed.sort(key=lambda x: x["order"])

    counters = {"essential": 0, "optional": 0, "custom": 0}
    catalog = []
    for item in parsed:
        cat = item["category"]
        idx = counters[cat]
        counters[cat] += 1
        catalog.append({
            "id": f"{cat}_{idx}",
            "category": cat,
            "check_id": item["check_id"],
            "command": item["command"],
            "description": item["description"] or "(설명 없음)",
            "enabled": item["enabled"],
        })
    return catalog


if __name__ == "__main__":
    catalog = load_catalog()
    print(f"카탈로그 항목 {len(catalog)}개 (필수 {sum(1 for c in catalog if c['category']=='essential')}, "
          f"선택 {sum(1 for c in catalog if c['category']=='optional')}, "
          f"커스텀 {sum(1 for c in catalog if c['category']=='custom')})")
    print("\n활성화된 커맨드:")
    for cmd in get_enabled_commands(catalog):
        print(" -", cmd)
